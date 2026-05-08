"""MeetPod 테이블 스펙 Excel 생성."""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter

NAVY = "0F1B3D"
BRAND = "5B7CFA"
ORANGE = "FF6900"
SUCCESS = "10B981"
DANGER = "EF4444"
WARN = "F59E0B"
INK = "0F172A"
MUTED = "94A3B8"
LIGHT = "F1F5F9"
HAIR = "E2E8F0"
WHITE = "FFFFFF"
HEADER_BG = "1F2A4D"

THIN = Side(style="thin", color=HAIR)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_title(cell, text, color=NAVY, size=16):
    cell.value = text
    cell.font = Font(name="맑은 고딕", size=size, bold=True,
                     color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(vertical="center", indent=1)


def style_section(cell, text, color=BRAND):
    cell.value = text
    cell.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(vertical="center", indent=1)


def style_header(cell, text, bg=NAVY):
    cell.value = text
    cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_cell(cell, text, bold=False, color=INK, bg=None, align="left",
               wrap=True):
    cell.value = text
    cell.font = Font(name="맑은 고딕", size=10, bold=bold, color=color)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=wrap)
    cell.border = BORDER


def merge_title(ws, row, text, span=6, color=NAVY, size=16, height=28):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=span)
    style_title(ws.cell(row=row, column=1), text, color, size)
    ws.row_dimensions[row].height = height


def merge_section(ws, row, text, span=6, color=BRAND, height=22):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=span)
    style_section(ws.cell(row=row, column=1), text, color)
    ws.row_dimensions[row].height = height


def write_meta(ws, row, items):
    """[(label, value), ...] 메타 행 작성"""
    for i, (k, v) in enumerate(items):
        r = row + i
        style_cell(ws.cell(row=r, column=1), k, bold=True, bg=LIGHT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        style_cell(ws.cell(row=r, column=2), v)
    return row + len(items)


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ===== Sheet: Overview =====
ws = wb.create_sheet("00_Overview")
set_column_widths(ws, [4, 24, 14, 60, 14, 18])
merge_title(ws, 1, "MeetPod — Table Specification", color=ORANGE)
write_meta(ws, 2, [
    ("Date", "2026-05-03"),
    ("DB", "Supabase (PostgreSQL 15)"),
    ("Status", "Initial spec (Phase 1~4 MVP)"),
    ("Companion", "system_design.md, MeetPod_화면설계.pptx"),
])

merge_section(ws, 7, "Tables")
hdr_row = 8
headers = ["#", "Table", "PK", "Description", "Sensitive", "RLS"]
for i, h in enumerate(headers, 1):
    style_header(ws.cell(row=hdr_row, column=i), h, NAVY)

tables_overview = [
    ("01", "profiles", "id (uuid)",
     "사용자 프로필 (auth.users 1:1)", "Y", "self / 같은 그룹·약속"),
    ("02", "friendships", "(a, b)",
     "친구 관계 (a<b 정규화)", "Y", "본인 포함만"),
    ("03", "invites", "code",
     "친구·그룹 초대 코드", "N", "anon 조회 (만료 체크)"),
    ("04", "groups", "id",
     "영구 그룹 (owner + admins + members)", "N", "멤버만"),
    ("05", "group_members", "(group, user)",
     "그룹 멤버십 + 역할", "N", "같은 그룹 멤버만"),
    ("06", "meetups", "id",
     "약속 (1회성/그룹)", "N", "참여자만"),
    ("07", "meetup_participants", "(meetup, user)",
     "약속 참여자 + 본인 위치공유 토글", "N", "같은 약속 참여자만"),
    ("08", "meetup_reminders", "(meetup, user, mins)",
     "사용자별 알림 큐 (발송 후 삭제)", "Y", "본인만"),
    ("09", "chat_rooms", "id",
     "채팅방 (그룹 또는 약속)", "N", "멤버십 체크"),
    ("10", "messages", "id",
     "채팅 메시지 (text/image/place/system)", "Y", "방 멤버만"),
    ("11", "location_pings", "id (bigserial)",
     "위치 핑 (TTL 24h)", "Y (HIGH)", "약속 참여자만, 본인 INSERT"),
    ("12", "push_outbox", "id",
     "푸시 idempotent 큐 (Phase 2)", "N", "service only"),
]
for i, row in enumerate(tables_overview):
    r = hdr_row + 1 + i
    for j, v in enumerate(row, 1):
        bg = WHITE if i % 2 == 0 else LIGHT
        style_cell(ws.cell(row=r, column=j), v, bg=bg)

start = hdr_row + 1 + len(tables_overview) + 2
merge_section(ws, start, "Conventions", color=SUCCESS)
conv = [
    ("ID", "uuid PK DEFAULT gen_random_uuid() (예외: auth.users.id, invites.code)"),
    ("시간", "timestamptz 통일. DEFAULT now()"),
    ("삭제", "기본 hard delete. 사용자 콘텐츠(messages)만 soft (deleted_at)"),
    ("외래 키", "ON DELETE CASCADE 기본, 사용자 참조는 RESTRICT"),
    ("명명", "테이블/컬럼 snake_case. boolean은 is_*/has_* 접두"),
    ("인덱스", "idx_<table>_<col>(_col2...) 명명"),
    ("트리거", "trg_<table>_<event> 명명"),
    ("RLS", "모든 public 테이블 ENABLE. 정책명 <verb>_<who>"),
    ("Extensions", "pgcrypto, pg_cron"),
]
write_meta(ws, start + 1, conv)


# ─── Helper to build a table sheet ───────────────────────
def build_table_sheet(idx, name, description,
                      columns, constraints, indexes, rls, triggers=None,
                      notes=None):
    ws = wb.create_sheet(f"{idx:02d}_{name}")
    set_column_widths(ws, [22, 22, 8, 14, 18, 60])
    merge_title(ws, 1, f"{idx:02d}.  {name}", color=ORANGE, size=15)
    write_meta(ws, 2, [("Description", description)])

    row = 4
    merge_section(ws, row, "Columns", color=BRAND)
    row += 1
    headers = ["Column", "Type", "Null", "Default", "PK/FK", "Description"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(row=row, column=i), h, NAVY)
    row += 1
    for i, col in enumerate(columns):
        for j, v in enumerate(col, 1):
            bg = WHITE if i % 2 == 0 else LIGHT
            bold = (j == 1)
            style_cell(ws.cell(row=row, column=j), v, bold=bold, bg=bg)
        ws.row_dimensions[row].height = 30
        row += 1

    if constraints:
        row += 1
        merge_section(ws, row, "Constraints", color=DANGER)
        row += 1
        for c in constraints:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            style_cell(ws.cell(row=row, column=1), c, bg=WHITE)
            row += 1

    if indexes:
        row += 1
        merge_section(ws, row, "Indexes", color=WARN)
        row += 1
        for c in indexes:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            style_cell(ws.cell(row=row, column=1), c, bg=WHITE)
            row += 1

    row += 1
    merge_section(ws, row, "RLS Policies", color=SUCCESS)
    row += 1
    headers = ["Operation", "Policy"]
    style_header(ws.cell(row=row, column=1), "Operation", NAVY)
    ws.merge_cells(start_row=row, start_column=2,
                   end_row=row, end_column=6)
    style_header(ws.cell(row=row, column=2), "Policy", NAVY)
    row += 1
    for i, (op, policy) in enumerate(rls):
        bg = WHITE if i % 2 == 0 else LIGHT
        style_cell(ws.cell(row=row, column=1), op, bold=True, bg=bg,
                   align="center")
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=6)
        style_cell(ws.cell(row=row, column=2), policy, bg=bg)
        ws.row_dimensions[row].height = 32
        row += 1

    if triggers:
        row += 1
        merge_section(ws, row, "Triggers / Functions", color="8B5CF6")
        row += 1
        for t in triggers:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            style_cell(ws.cell(row=row, column=1), t, bg=WHITE)
            ws.row_dimensions[row].height = 28
            row += 1

    if notes:
        row += 1
        merge_section(ws, row, "Notes", color=MUTED)
        row += 1
        for n in notes:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            style_cell(ws.cell(row=row, column=1), n, bg=WHITE)
            ws.row_dimensions[row].height = 28
            row += 1


# ===== 01_profiles =====
build_table_sheet(
    1, "profiles",
    "auth.users와 1:1. 가입 직후 트리거로 행 생성, "
    "핸들은 OnboardingHandleScreen에서 채움.",
    [
        ("id", "uuid", "NO", "—",
         "PK, FK→auth.users.id (CASCADE)", "사용자 ID"),
        ("handle", "text", "NO", "—", "UNIQUE",
         "@handle. 소문자 영숫자·언더스코어. 4~20자"),
        ("display_name", "text", "NO", "—", "—",
         "화면 표시 이름. 1~30자"),
        ("avatar_url", "text", "YES", "NULL", "—",
         "Supabase Storage URL"),
        ("expo_push_token", "text", "YES", "NULL", "—",
         "푸시 발송용"),
        ("default_reminder_minutes", "int", "NO", "30", "—",
         "새 약속의 기본 개인 알림 (분)"),
        ("locale", "text", "NO", "'ko'", "—", "ko / en"),
        ("created_at", "timestamptz", "NO", "now()", "—", ""),
        ("updated_at", "timestamptz", "NO", "now()", "—",
         "트리거로 자동 갱신"),
    ],
    [
        "CHECK (handle ~ '^[a-z0-9_]{4,20}$')",
        "CHECK (char_length(display_name) BETWEEN 1 AND 30)",
    ],
    ["PK on (id)", "UNIQUE on (handle)"],
    [
        ("SELECT", "auth.uid() = id OR EXISTS(같은 group_members) "
                   "OR EXISTS(같은 meetup_participants)"),
        ("UPDATE", "auth.uid() = id (id/created_at 변경 금지)"),
        ("INSERT/DELETE", "트리거 외 차단"),
    ],
    triggers=[
        "trg_profiles_updated_at: BEFORE UPDATE → updated_at = now()",
        "trg_users_after_insert (auth.users): INSERT new profile "
        "with display_name='사용자', handle=NULL",
    ],
)

# ===== 02_friendships =====
build_table_sheet(
    2, "friendships",
    "A↔B 친구 관계. 정규화로 양방향 한 행만 저장.",
    [
        ("user_a_id", "uuid", "NO", "—",
         "FK→profiles.id (CASCADE)", "작은 쪽 ID"),
        ("user_b_id", "uuid", "NO", "—",
         "FK→profiles.id (CASCADE)", "큰 쪽 ID"),
        ("created_at", "timestamptz", "NO", "now()", "—", ""),
    ],
    [
        "PRIMARY KEY (user_a_id, user_b_id)",
        "CHECK (user_a_id < user_b_id)",
    ],
    [
        "PK on (user_a_id, user_b_id)",
        "idx_friendships_b on (user_b_id, user_a_id)  -- 역방향 조회",
    ],
    [
        ("SELECT", "auth.uid() IN (user_a_id, user_b_id)"),
        ("INSERT/DELETE", "backend service role 전용 (invite_service)"),
    ],
    triggers=[
        "Helper SQL: upsert_friendship(u1, u2) — "
        "LEAST/GREATEST 정규화 + ON CONFLICT DO NOTHING",
    ],
)

# ===== 03_invites =====
build_table_sheet(
    3, "invites",
    "친구 또는 그룹 초대 코드. 8자 Crockford base32.",
    [
        ("code", "text", "NO", "—", "PK", "8자 URL-safe"),
        ("inviter_id", "uuid", "NO", "—",
         "FK→profiles.id", "발급자"),
        ("kind", "text", "NO", "—",
         "CHECK ('friend','group')", ""),
        ("target_group_id", "uuid", "YES", "NULL",
         "FK→groups.id (CASCADE)", "kind='group'일 때 필수"),
        ("expires_at", "timestamptz", "NO", "—", "—",
         "발급 후 7일 기본"),
        ("max_uses", "int", "NO", "10", "—", ""),
        ("used_count", "int", "NO", "0", "—", ""),
        ("created_at", "timestamptz", "NO", "now()", "—", ""),
        ("revoked_at", "timestamptz", "YES", "NULL", "—",
         "발급자 취소 시"),
    ],
    [
        "CHECK (code ~ '^[A-HJ-NP-Z0-9]{8}$')",
        "CHECK ((kind='group' AND target_group_id IS NOT NULL) "
        "OR (kind='friend' AND target_group_id IS NULL))",
        "CHECK (used_count <= max_uses)",
    ],
    [
        "PK on (code)",
        "idx_invites_inviter on (inviter_id, created_at DESC)",
        "idx_invites_group on (target_group_id) WHERE kind='group'",
    ],
    [
        ("SELECT", "anonymous OK. 단, expires_at > now() "
                   "AND revoked_at IS NULL 정책"),
        ("INSERT", "auth.uid() = inviter_id AND "
                   "(kind='friend' OR group owner/admin)"),
        ("UPDATE/DELETE", "auth.uid() = inviter_id (revoke)"),
    ],
    triggers=[
        "RPC: accept_invite(p_code) — used_count++ + "
        "friendships/group_members INSERT 트랜잭션 (SECURITY DEFINER)",
    ],
)

# ===== 04_groups =====
build_table_sheet(
    4, "groups",
    "영구 그룹. owner 1명 + admins. 누구나 생성 가능.",
    [
        ("id", "uuid", "NO", "gen_random_uuid()", "PK", ""),
        ("name", "text", "NO", "—", "—", "1~40자"),
        ("description", "text", "YES", "NULL", "—", "0~200자"),
        ("avatar_url", "text", "YES", "NULL", "—", ""),
        ("owner_id", "uuid", "NO", "—",
         "FK→profiles.id (RESTRICT)", "위임 시 변경"),
        ("created_at", "timestamptz", "NO", "now()", "—", ""),
        ("updated_at", "timestamptz", "NO", "now()", "—", "트리거"),
    ],
    [
        "CHECK (char_length(name) BETWEEN 1 AND 40)",
        "CHECK (description IS NULL OR char_length(description) <= 200)",
    ],
    ["PK on (id)", "idx_groups_owner on (owner_id)"],
    [
        ("SELECT", "EXISTS(group_members WHERE group_id=id "
                   "AND user_id=auth.uid())"),
        ("INSERT", "auth.uid() = owner_id (트리거가 owner를 "
                   "group_members에 자동 추가)"),
        ("UPDATE (info)", "owner 또는 admin"),
        ("UPDATE (owner)", "현재 owner만 (위임)"),
        ("DELETE", "owner만"),
    ],
    triggers=[
        "trg_groups_after_insert: group_members(role='owner') INSERT "
        "+ chat_rooms(kind='group', ref_id=NEW.id) INSERT",
    ],
)

# ===== 05_group_members =====
build_table_sheet(
    5, "group_members",
    "그룹 멤버십과 역할.",
    [
        ("group_id", "uuid", "NO", "—",
         "FK→groups.id (CASCADE)", ""),
        ("user_id", "uuid", "NO", "—",
         "FK→profiles.id (CASCADE)", ""),
        ("role", "text", "NO", "—",
         "CHECK ('owner','admin','member')", ""),
        ("joined_at", "timestamptz", "NO", "now()", "—", ""),
    ],
    ["PRIMARY KEY (group_id, user_id)"],
    [
        "PK on (group_id, user_id)",
        "idx_group_members_user on (user_id, joined_at DESC) "
        "-- 내 그룹 목록",
    ],
    [
        ("SELECT", "본인 멤버인 그룹만 (EXISTS 같은 그룹)"),
        ("INSERT", "backend service (accept_invite RPC)"),
        ("UPDATE (role)", "owner만 (admin↔member). owner 위임은 "
                          "groups.owner_id 갱신 트랜잭션"),
        ("DELETE", "owner/admin (추방), 본인 (탈퇴). "
                   "마지막 owner 탈퇴 차단"),
    ],
    triggers=[
        "trg_group_members_prevent_owner_leave: "
        "owner role DELETE 차단 (먼저 위임 필요)",
    ],
)

# ===== 06_meetups =====
build_table_sheet(
    6, "meetups",
    "약속. 단발 또는 그룹. 핵심 엔티티.",
    [
        ("id", "uuid", "NO", "gen_random_uuid()", "PK", ""),
        ("group_id", "uuid", "YES", "NULL",
         "FK→groups.id (CASCADE)", "NULL이면 1회성"),
        ("creator_id", "uuid", "NO", "—",
         "FK→profiles.id", ""),
        ("title", "text", "NO", "—", "—", "1~60자"),
        ("starts_at", "timestamptz", "NO", "—", "—", ""),
        ("ends_at", "timestamptz", "NO", "—", "—", ""),
        ("place_name", "text", "NO", "—", "—", ""),
        ("place_lat", "double precision", "NO", "—", "—", ""),
        ("place_lng", "double precision", "NO", "—", "—", ""),
        ("place_address", "text", "YES", "NULL", "—", ""),
        ("place_google_id", "text", "YES", "NULL", "—",
         "Google Place ID"),
        ("location_share_minutes_before", "int", "NO", "20", "—",
         "0~240분"),
        ("status", "text", "NO", "'scheduled'", "—",
         "scheduled/active/ended/cancelled"),
        ("created_at", "timestamptz", "NO", "now()", "—", ""),
        ("updated_at", "timestamptz", "NO", "now()", "—", "트리거"),
    ],
    [
        "CHECK (ends_at > starts_at)",
        "CHECK (place_lat BETWEEN -90 AND 90)",
        "CHECK (place_lng BETWEEN -180 AND 180)",
        "CHECK (location_share_minutes_before BETWEEN 0 AND 240)",
        "CHECK (char_length(title) BETWEEN 1 AND 60)",
    ],
    [
        "PK on (id)",
        "idx_meetups_group_starts on (group_id, starts_at DESC) "
        "WHERE group_id IS NOT NULL",
        "idx_meetups_creator_starts on (creator_id, starts_at DESC)",
        "idx_meetups_status_ends on (status, ends_at) -- pg_cron 전환용",
    ],
    [
        ("SELECT", "EXISTS(meetup_participants WHERE meetup_id=id "
                   "AND user_id=auth.uid())"),
        ("INSERT", "auth.uid() = creator_id "
                   "(트리거가 participants/chat_room/reminder 생성)"),
        ("UPDATE", "creator 또는 (group_id의 owner/admin)"),
        ("DELETE", "위 UPDATE 권한자만"),
    ],
    triggers=[
        "trg_meetups_after_insert: meetup_participants(creator) + "
        "chat_rooms(kind='meetup') + meetup_reminders(creator, default) INSERT",
    ],
)

# ===== 07_meetup_participants =====
build_table_sheet(
    7, "meetup_participants",
    "약속 참여자 + 본인 위치공유 토글.",
    [
        ("meetup_id", "uuid", "NO", "—",
         "FK→meetups.id (CASCADE)", ""),
        ("user_id", "uuid", "NO", "—",
         "FK→profiles.id (CASCADE)", ""),
        ("status", "text", "NO", "'going'",
         "CHECK ('going')", "MVP는 going만. RSVP는 Phase 2"),
        ("share_location", "boolean", "NO", "true", "—",
         "본인 위치공유 ON/OFF (시나리오 9)"),
        ("joined_at", "timestamptz", "NO", "now()", "—", ""),
    ],
    ["PRIMARY KEY (meetup_id, user_id)"],
    [
        "PK on (meetup_id, user_id)",
        "idx_meetup_participants_user on (user_id, joined_at DESC) "
        "-- 내 약속 목록",
    ],
    [
        ("SELECT", "본인 참여 meetup의 참여자만 (EXISTS 같은 meetup)"),
        ("INSERT", "creator/admin가 추가, 본인은 자기 참여"),
        ("UPDATE (share_location)", "auth.uid() = user_id (본인만 토글)"),
        ("DELETE", "creator/admin가 추방 OR 본인 탈퇴"),
    ],
)

# ===== 08_meetup_reminders =====
build_table_sheet(
    8, "meetup_reminders",
    "사용자별 개인 알림 큐. 푸시 발송 후 행 삭제.",
    [
        ("meetup_id", "uuid", "NO", "—",
         "FK→meetups.id (CASCADE)", ""),
        ("user_id", "uuid", "NO", "—",
         "FK→profiles.id (CASCADE)", ""),
        ("minutes_before", "int", "NO", "—", "—", "0~10080 (7일)"),
        ("notify_at", "timestamptz", "NO", "—", "—",
         "starts_at - interval. 발송 큐 인덱스용"),
    ],
    ["PRIMARY KEY (meetup_id, user_id, minutes_before)"],
    [
        "PK on (meetup_id, user_id, minutes_before)",
        "idx_reminders_notify_at on (notify_at) — 1분 cron 큐",
    ],
    [
        ("SELECT/INSERT/UPDATE/DELETE", "auth.uid() = user_id"),
    ],
    triggers=[
        "trg_meetups_update_reminders (meetups UPDATE OF starts_at): "
        "관련 reminder의 notify_at 재계산",
        "Edge Function (1분 cron): SELECT WHERE notify_at <= now() → "
        "Expo push → DELETE",
    ],
)

# ===== 09_chat_rooms =====
build_table_sheet(
    9, "chat_rooms",
    "그룹 / 약속별 채팅방. (kind, ref_id) 1:1 매핑.",
    [
        ("id", "uuid", "NO", "gen_random_uuid()", "PK", ""),
        ("kind", "text", "NO", "—",
         "CHECK ('group','meetup')", ""),
        ("ref_id", "uuid", "NO", "—", "—",
         "groups.id 또는 meetups.id"),
        ("archived_at", "timestamptz", "YES", "NULL", "—",
         "약속 종료 시 세팅"),
        ("created_at", "timestamptz", "NO", "now()", "—", ""),
    ],
    ["UNIQUE (kind, ref_id)"],
    ["PK on (id)", "UNIQUE on (kind, ref_id)"],
    [
        ("SELECT (group)", "EXISTS(group_members WHERE group_id=ref_id "
                           "AND user_id=auth.uid())"),
        ("SELECT (meetup)", "EXISTS(meetup_participants WHERE "
                            "meetup_id=ref_id AND user_id=auth.uid())"),
        ("INSERT", "트리거 외 차단"),
        ("UPDATE (archived_at)", "pg_cron service 또는 약속 권한자"),
        ("DELETE", "차단 (메시지 보존)"),
    ],
)

# ===== 10_messages =====
build_table_sheet(
    10, "messages",
    "채팅 메시지. text/image/place/system. 본인만 편집/soft-delete.",
    [
        ("id", "uuid", "NO", "gen_random_uuid()", "PK", ""),
        ("room_id", "uuid", "NO", "—",
         "FK→chat_rooms.id (CASCADE)", ""),
        ("sender_id", "uuid", "NO", "—",
         "FK→profiles.id", ""),
        ("kind", "text", "NO", "—",
         "CHECK ('text','image','place','system')", ""),
        ("body", "text", "YES", "NULL", "—", "text/system 시 필수"),
        ("image_url", "text", "YES", "NULL", "—", "image 시 필수"),
        ("place_payload", "jsonb", "YES", "NULL", "—",
         "place 시 필수: {name,lat,lng,google_id,address,url}"),
        ("reply_to_id", "uuid", "YES", "NULL",
         "FK→messages.id (SET NULL)", "답장"),
        ("created_at", "timestamptz", "NO", "now()", "—", ""),
        ("edited_at", "timestamptz", "YES", "NULL", "—",
         "편집 시 갱신"),
        ("deleted_at", "timestamptz", "YES", "NULL", "—",
         "soft delete"),
    ],
    [
        "CHECK ((kind='text' AND body IS NOT NULL AND "
        "char_length(body) BETWEEN 1 AND 4000) "
        "OR (kind='image' AND image_url IS NOT NULL) "
        "OR (kind='place' AND place_payload IS NOT NULL) "
        "OR (kind='system' AND body IS NOT NULL))",
    ],
    [
        "PK on (id)",
        "idx_messages_room_created on (room_id, created_at DESC) "
        "-- 채팅 페이지네이션",
        "idx_messages_sender on (sender_id, created_at DESC)",
    ],
    [
        ("SELECT", "chat_rooms RLS와 동일 멤버십 로직"),
        ("INSERT", "위 + auth.uid() = sender_id (archived 체크는 "
                   "API에서, 시나리오 11)"),
        ("UPDATE", "auth.uid() = sender_id (body/edited/deleted만)"),
        ("DELETE", "차단 (soft delete만)"),
    ],
    notes=[
        "Realtime publication: ALTER PUBLICATION supabase_realtime "
        "ADD TABLE messages;",
        "모바일이 room_id 필터로 INSERT/UPDATE 구독",
    ],
)

# ===== 11_location_pings =====
build_table_sheet(
    11, "location_pings",
    "위치 핑. 짧은 TTL (24h). 약속 active 상태에서만 INSERT.",
    [
        ("id", "bigserial", "NO", "—", "PK", "uuid 비용 절감"),
        ("meetup_id", "uuid", "NO", "—",
         "FK→meetups.id (CASCADE)", ""),
        ("user_id", "uuid", "NO", "—",
         "FK→profiles.id (CASCADE)", ""),
        ("lat", "double precision", "NO", "—",
         "CHECK (-90~90)", ""),
        ("lng", "double precision", "NO", "—",
         "CHECK (-180~180)", ""),
        ("accuracy_m", "real", "YES", "NULL", "—", "미터 단위"),
        ("recorded_at", "timestamptz", "NO", "now()", "—",
         "클라이언트 타임스탬프"),
    ],
    [
        "CHECK (lat BETWEEN -90 AND 90)",
        "CHECK (lng BETWEEN -180 AND 180)",
    ],
    [
        "PK on (id)",
        "idx_pings_meetup_recorded on (meetup_id, recorded_at DESC)",
        "idx_pings_meetup_user_recorded on "
        "(meetup_id, user_id, recorded_at DESC) -- 멤버 trail",
    ],
    [
        ("SELECT", "같은 meetup 참여자 (EXISTS meetup_participants)"),
        ("INSERT", "auth.uid() = user_id AND 본인 참여 AND "
                   "share_location=true AND meetups.status='active'"),
        ("UPDATE/DELETE", "차단 (cron만 정리)"),
    ],
    notes=[
        "Realtime: 모바일이 meetup_id 필터로 INSERT 구독",
        "멤버별 최신 핀만 지도에 표시 (클라이언트 dedup)",
        "pg_cron cleanup_location_pings: ends_at + 24h 지난 핑 DELETE",
    ],
)

# ===== 12_push_outbox =====
build_table_sheet(
    12, "push_outbox",
    "푸시 idempotent 큐 (Phase 2). MVP에서는 reminder 직접 DELETE.",
    [
        ("id", "uuid", "NO", "gen_random_uuid()", "PK", ""),
        ("user_id", "uuid", "NO", "—", "FK→profiles.id", ""),
        ("kind", "text", "NO", "—", "—",
         "reminder/invite/meetup_change/cancel"),
        ("payload", "jsonb", "NO", "—", "—", "푸시 메시지 데이터"),
        ("dedupe_key", "text", "NO", "—", "UNIQUE",
         "예: reminder:{meetup}:{user}:{minutes}"),
        ("sent_at", "timestamptz", "YES", "NULL", "—", ""),
        ("created_at", "timestamptz", "NO", "now()", "—", ""),
    ],
    ["UNIQUE (dedupe_key)"],
    [
        "PK on (id)",
        "UNIQUE on (dedupe_key)",
        "idx_push_outbox_pending on (sent_at) WHERE sent_at IS NULL",
    ],
    [
        ("SELECT/INSERT/UPDATE/DELETE", "service role only"),
    ],
)


# ===== 13_pg_cron =====
ws = wb.create_sheet("13_pg_cron")
set_column_widths(ws, [4, 26, 12, 60, 50])
merge_title(ws, 1, "13.  pg_cron 작업", color=ORANGE, span=5)
write_meta(ws, 2, [
    ("Description",
     "Supabase pg_cron으로 1분~1일 주기 작업. Edge Function "
     "또는 SQL로 구현."),
])
row = 5
merge_section(ws, row, "Cron Jobs", color=BRAND, span=5)
row += 1
for i, h in enumerate(["#", "Name", "Schedule", "작업", "비고"], 1):
    style_header(ws.cell(row=row, column=i), h, NAVY)
row += 1
crons = [
    ("1", "tick_meetup_status", "*/1 * * * *",
     "scheduled→active (starts_at <= now), active→ended "
     "(ends_at <= now) + 해당 chat_rooms.archived_at = now()",
     "클라이언트 시간 조작 무관"),
    ("2", "dispatch_reminders", "*/1 * * * *",
     "meetup_reminders WHERE notify_at <= now → Expo push 발송 후 DELETE",
     "Edge Function (push_service 호출)"),
    ("3", "cleanup_location_pings", "0 * * * *",
     "ends_at + 24h 지난 meetup의 핑 DELETE",
     "프라이버시 + 저장 비용"),
    ("4", "cleanup_expired_invites", "0 0 * * *",
     "expires_at 지난 used_count=0 invites DELETE",
     "선택, MVP에서 생략 가능"),
]
for i, c in enumerate(crons):
    bg = WHITE if i % 2 == 0 else LIGHT
    for j, v in enumerate(c, 1):
        bold = (j == 2)
        style_cell(ws.cell(row=row, column=j), v, bold=bold, bg=bg)
    ws.row_dimensions[row].height = 50
    row += 1


# ===== 14_storage =====
ws = wb.create_sheet("14_Storage")
set_column_widths(ws, [4, 22, 80, 30])
merge_title(ws, 1, "14.  Supabase Storage", color=ORANGE, span=4)
row = 3
merge_section(ws, row, "Buckets", color=BRAND, span=4)
row += 1
for i, h in enumerate(["#", "Bucket", "Policy", "경로 규칙"], 1):
    style_header(ws.cell(row=row, column=i), h, NAVY)
row += 1
buckets = [
    ("1", "chat-images",
     "INSERT: 인증 사용자. SELECT: 메시지의 chat_room 멤버만 "
     "(백엔드 signed URL 권장 또는 storage policies)",
     "chat-images/{room_id}/{uuid}.jpg"),
    ("2", "avatars",
     "INSERT: 본인 폴더만. SELECT: 모든 인증 사용자 (검색 노출 X)",
     "avatars/{user_id}/{uuid}.jpg"),
]
for i, b in enumerate(buckets):
    bg = WHITE if i % 2 == 0 else LIGHT
    for j, v in enumerate(b, 1):
        bold = (j == 2)
        style_cell(ws.cell(row=row, column=j), v, bold=bold, bg=bg)
    ws.row_dimensions[row].height = 60
    row += 1


# ===== 15_migrations =====
ws = wb.create_sheet("15_Migrations")
set_column_widths(ws, [4, 36, 80])
merge_title(ws, 1, "15.  Migration 순서", color=ORANGE, span=3)
row = 3
merge_section(ws, row, "Files", color=BRAND, span=3)
row += 1
for i, h in enumerate(["#", "File", "내용"], 1):
    style_header(ws.cell(row=row, column=i), h, NAVY)
row += 1
migs = [
    ("001", "001_init_extensions.sql", "pgcrypto, pg_cron 활성화"),
    ("002", "002_profiles.sql",
     "profiles + auth.users 트리거 + RLS"),
    ("003", "003_friendships.sql",
     "friendships + helper RPC + RLS"),
    ("004", "004_groups.sql",
     "groups + group_members + 트리거 + RLS"),
    ("005", "005_invites.sql",
     "invites + accept_invite RPC + RLS"),
    ("006", "006_meetups.sql",
     "meetups + meetup_participants + 트리거 + RLS"),
    ("007", "007_reminders.sql",
     "meetup_reminders + 재계산 트리거 + RLS"),
    ("008", "008_chat.sql",
     "chat_rooms + messages + Realtime publication + RLS"),
    ("009", "009_location.sql",
     "location_pings + cron 등록 + RLS"),
    ("010", "010_storage.sql",
     "chat-images / avatars 버킷 + storage policies"),
]
for i, m in enumerate(migs):
    bg = WHITE if i % 2 == 0 else LIGHT
    for j, v in enumerate(m, 1):
        bold = (j == 2)
        style_cell(ws.cell(row=row, column=j), v, bold=bold, bg=bg)
    row += 1


# ===== Save =====
out = "d:/Workspace/CPKWorks/MeetPod/docs/MeetPod_테이블스펙.xlsx"
wb.save(out)
